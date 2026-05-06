# File upload router: preview, ingest CSV/Excel into etl_raw, and download
# datasource-specific data-entry templates (Excel with 3 instruction sheets, or CSV).
# Template columns come from the datasource's src_field mappings so the file exactly
# matches what the system expects — no manual column naming needed.
import io
import os
import shutil
import uuid
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from database import get_db
from models import DataSource, FieldMapping
from models.data_source import SourceType
from routers.auth import get_current_user
from services.ingest_service import ingest_file, preview_file

router = APIRouter(prefix="/api/v1/upload", tags=["upload"])

TMP_DIR = "/app/tmp"
ALLOWED_EXT = {".csv", ".xlsx", ".xls"}


def _save_tmp(upload: UploadFile) -> str:
    suffix = Path(upload.filename or "file").suffix.lower()
    tmp_path = os.path.join(TMP_DIR, f"{uuid.uuid4()}{suffix}")
    with open(tmp_path, "wb") as f:
        shutil.copyfileobj(upload.file, f)
    return tmp_path


def _check_ext(filename: str) -> None:
    if Path(filename).suffix.lower() not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail="仅支持 CSV / Excel (.xlsx/.xls) 文件")


@router.post("/preview")
async def preview(
    file: UploadFile = File(...),
    _=Depends(get_current_user),
):
    _check_ext(file.filename or "")
    tmp_path = _save_tmp(file)
    try:
        return preview_file(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"文件解析失败: {str(e)}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/{ds_id}")
async def upload_and_ingest(
    ds_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    _check_ext(file.filename or "")

    ds = db.get(DataSource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="数据源不存在")
    if ds.source_type != SourceType.upload:
        raise HTTPException(status_code=400, detail="该数据源不是文件上传类型")

    tmp_path = _save_tmp(file)
    try:
        return ingest_file(tmp_path, file.filename or "unknown", ds.target_raw_table)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"数据入库失败: {str(e)}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ── Upload template ───────────────────────────────────────────────────────────

# Sample values shown in the template's example row, keyed by dst_type value.
_SAMPLE = {
    "string":   "示例文本",
    "integer":  "42",
    "float":    "123.45",
    "date":     "2024-01-15",
    "datetime": "2024-01-15 09:30:00",
    "boolean":  "TRUE",
}

_TYPE_FMT = {
    "string":   "任意文本",
    "integer":  "整数，不含小数点和千位分隔符（如 42，不是 42.0 或 42,000）",
    "float":    "小数，使用英文小数点（如 123.45，不是 123,45）",
    "date":     "日期，格式必须为 YYYY-MM-DD（如 2024-01-15，不是 2024/1/15）",
    "datetime": "日期时间，格式必须为 YYYY-MM-DD HH:MM:SS（如 2024-01-15 09:30:00）",
    "boolean":  "布尔值，填写 TRUE 或 FALSE（英文大写，不支持 是/否 或 1/0）",
}

_FAQ = [
    ("上传后数据进入 Raw 层，但执行 Raw→DWD 后某些字段为空（NULL）？",
     "Raw 层将所有数据原样存为文本；类型转换在执行 Raw→DWD 时进行。\n"
     "请检查：① 字段值是否符合格式要求（见「字段说明」页）；② 数字是否含千位分隔符或货币符号；③ 日期格式是否为 YYYY-MM-DD。"),
    ("上传成功，但执行 Raw→DWD 后行数少于上传行数？",
     "这是正常的 —— 过滤规则（FilterRule）会排除不满足条件的行。\n"
     "Raw 层始终保留全部原始数据；DWD 层是过滤后的结果。\n"
     "可在「执行历史」页查看每次执行的成功行数和失败行数。"),
    ("上传时提示「仅支持 CSV / Excel 文件」？",
     "系统只接受 .csv、.xlsx、.xls 格式。\n"
     "请勿将其他格式的文件（如 .txt、.json）改后缀名后上传。"),
    ("Excel 文件上传后字段名不正确，或第一列变成了序号？",
     "系统以文件第一行作为字段名（表头行）。请确保：\n"
     "① 第一行是字段名，不是合并单元格的大标题；\n"
     "② 没有多余的空白行在表头之前；\n"
     "③ 删除本模板的示例行（第 2 行）再上传。"),
    ("Excel 文件有多个 Sheet，数据只被部分读取？",
     "系统只读取第一个 Sheet 中的数据。\n"
     "请将所有数据整理到第一个 Sheet（本模板的「数据录入」页），其余 Sheet 不会被读取。"),
    ("上传数据后，如何确认数据写入正确？",
     "上传成功后可通过以下方式确认：\n"
     "① 上传页面会显示「入库结果」，含入库行数和前 5 行数据预览；\n"
     "② 在「数据预览」页选择该数据源 → RAW 层，可浏览完整原始数据；\n"
     "③ 执行 Raw→DWD 后，在「数据预览」页切换到 DWD 层查看转换结果。"),
]


@router.get("/{ds_id}/template")
def download_upload_template(
    ds_id: int,
    format: str = Query(default="excel", pattern="^(csv|excel)$"),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    ds = db.get(DataSource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="数据源不存在")

    mappings = (
        db.query(FieldMapping)
        .filter(FieldMapping.data_source_id == ds_id, FieldMapping.skip == False)
        .order_by(FieldMapping.sort_order, FieldMapping.id)
        .all()
    )
    if not mappings:
        raise HTTPException(
            status_code=404,
            detail="该数据源尚未配置字段映射，请先在「字段映射」页保存映射规则后再下载模板",
        )

    fields = [(m.src_field, m.dst_type.value) for m in mappings]

    if format == "csv":
        return _csv_template(ds, fields)
    return _excel_template(ds, fields)


def _csv_template(ds: DataSource, fields: list[tuple[str, str]]) -> StreamingResponse:
    headers = [f for f, _ in fields]
    sample = [_SAMPLE.get(t, "") for _, t in fields]
    lines = [
        ",".join(headers),
        ",".join(f'"{v}"' if "," in v else v for v in sample),
    ]
    content = "\n".join(lines) + "\n"
    fname = f"upload_template_{ds.target_raw_table}.csv"
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),  # utf-8-sig so Excel opens correctly
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _excel_template(ds: DataSource, fields: list[tuple[str, str]]) -> StreamingResponse:
    wb = openpyxl.Workbook()

    # ── Sheet 1: 数据录入 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "数据录入"

    hdr_fill = PatternFill(fill_type="solid", fgColor="2F75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    sample_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    sample_font = Font(italic=True, color="7F7F7F", size=10)

    # Row 1: column headers
    for col_idx, (field, _) in enumerate(fields, 1):
        cell = ws1.cell(row=1, column=col_idx, value=field)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: sample values (yellow, italic) — user deletes this row before uploading
    for col_idx, (_, dtype) in enumerate(fields, 1):
        cell = ws1.cell(row=2, column=col_idx, value=_SAMPLE.get(dtype, ""))
        cell.font = sample_font
        cell.fill = sample_fill

    # Note in the column after the last field reminding user to delete the sample row
    note_col = len(fields) + 1
    note_cell = ws1.cell(row=2, column=note_col, value="← 示例行，上传前请删除本行")
    note_cell.font = Font(italic=True, color="FF0000", size=9)

    # Row 3: hint where real data starts
    ws1.cell(row=3, column=1, value="← 从本行开始填写真实数据").font = Font(
        color="AAAAAA", size=9, italic=True
    )

    for col_idx, (field, _) in enumerate(fields, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            len(field) + 4, 16
        )
    ws1.column_dimensions[openpyxl.utils.get_column_letter(note_col)].width = 28
    ws1.row_dimensions[1].height = 22

    # ── Sheet 2: 字段说明 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("字段说明")
    ws2.append(["字段名", "数据类型", "格式要求", "示例值"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
        cell.alignment = Alignment(wrap_text=True)

    for field, dtype in fields:
        ws2.append([field, dtype, _TYPE_FMT.get(dtype, ""), _SAMPLE.get(dtype, "")])

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 22

    for row in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    ws2.append([])
    note_row = ws2.append(["上传注意事项"]) or ws2[ws2.max_row]
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for line in [
        "• 所有字段值以文本形式上传，系统在执行 Raw→DWD 时自动进行类型转换",
        "• 类型转换失败的单元格会变为 NULL（空值），不会导致整行上传失败",
        "• 上传前请删除「数据录入」页的示例行（第 2 行）",
        "• 文件第一行必须是字段名行，不要在表头前添加额外标题行",
    ]:
        ws2.append([line])

    # ── Sheet 3: 常见问题 ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("常见问题")
    ws3.column_dimensions["A"].width = 75

    title_cell = ws3.cell(row=1, column=1,
                          value=f"数据上传常见问题与解决方法（数据源：{ds.name}）")
    title_cell.font = Font(bold=True, size=12)
    ws3.row_dimensions[1].height = 24
    ws3.append([])

    q_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    for i, (q, a) in enumerate(_FAQ, 1):
        q_cell = ws3.cell(row=ws3.max_row + 1, column=1, value=f"Q{i}：{q}")
        q_cell.font = Font(bold=True, color="1F4E79")
        q_cell.fill = q_fill
        q_cell.alignment = Alignment(wrap_text=True)
        ws3.row_dimensions[q_cell.row].height = 30

        a_cell = ws3.cell(row=ws3.max_row + 1, column=1, value=f"A：{a}")
        a_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[a_cell.row].height = max(a.count("\n") * 18 + 22, 44)
        ws3.append([])  # spacer between Q&A pairs

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"upload_template_{ds.target_raw_table}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
