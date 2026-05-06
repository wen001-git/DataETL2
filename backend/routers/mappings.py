# Field mapping configuration: how raw columns map to DWD columns (rename, type, default).
# Also exposes column-name helper endpoints used by downstream config pages:
#   GET /{ds_id}/raw-columns  → columns currently in etl_raw (metadata cols excluded)
#   GET /{ds_id}/dwd-columns  → columns in the configured etl_dwd table (post-execution)
#   GET /{ds_id}/dws-columns  → columns in the configured etl_dws table (post-execution)
# The dws-columns endpoint resolves the DWS table name via the datasource's agg_rule
# rather than accepting a table name from the caller, ensuring it always refers to the
# table that was actually written by the last DWD→DWS execution.
import io
import os
import shutil
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from database import engine, get_db
from models import DataSource, FieldMapping
from models.agg_rule import AggRule
from models.field_mapping import DstType
from models.mapping_version import MappingVersion
from routers.auth import get_current_user

router = APIRouter(prefix="/api/v1/datasources", tags=["mappings"])

DST_TYPES = ["string", "integer", "float", "date", "datetime", "boolean"]


# ── Schemas ──────────────────────────────────────────────────────────────────

class MappingItem(BaseModel):
    src_field: str
    dst_field: str
    dst_type: DstType = DstType.string
    default_value: Optional[str] = None
    skip: bool = False
    sort_order: int = 0


class MappingOut(BaseModel):
    id: int
    data_source_id: int
    src_field: str
    dst_field: str
    dst_type: DstType
    default_value: Optional[str]
    skip: bool
    sort_order: int
    target_dwd_table: str

    model_config = {"from_attributes": True}


class BulkSaveRequest(BaseModel):
    target_dwd_table: str
    mappings: list[MappingItem]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_ds(ds_id: int, db: Session) -> DataSource:
    ds = db.get(DataSource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="数据源不存在")
    return ds


_RAW_META = {"_src_file", "_ingested_at", "_run_id"}


def _schema_columns(schema: str, table: str, exclude: set[str] | None = None) -> list[str]:
    # Returns [] (not raises) if the table doesn't exist yet — callers convert that to 404.
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(f"SHOW COLUMNS FROM `{schema}`.`{table}`")
            ).fetchall()
        cols = [r[0] for r in rows]
        return [c for c in cols if c not in (exclude or set())]
    except Exception:
        return []


def _raw_columns(table: str) -> list[str]:
    return _schema_columns("etl_raw", table, exclude=_RAW_META)


def _dwd_columns(table: str) -> list[str]:
    return _schema_columns("etl_dwd", table)


def _dws_columns(table: str) -> list[str]:
    return _schema_columns("etl_dws", table)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/{ds_id}/dws-columns")
def get_dws_columns(
    ds_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    _get_ds(ds_id, db)
    agg_rule = db.query(AggRule).filter(AggRule.data_source_id == ds_id).first()
    if not agg_rule:
        raise HTTPException(status_code=404, detail="没有聚合规则，请先配置 DWS 聚合规则")
    cols = _dws_columns(agg_rule.target_dws_table)
    if not cols:
        raise HTTPException(
            status_code=404,
            detail=f"DWS 表 '{agg_rule.target_dws_table}' 不存在，请先执行 DWD→DWS",
        )
    return {"columns": cols, "dws_table": agg_rule.target_dws_table}


@router.get("/{ds_id}/dwd-columns")
def get_dwd_columns(
    ds_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    _get_ds(ds_id, db)
    first_mapping = (
        db.query(FieldMapping)
        .filter(FieldMapping.data_source_id == ds_id, FieldMapping.skip == False)
        .order_by(FieldMapping.sort_order)
        .first()
    )
    if not first_mapping:
        raise HTTPException(status_code=404, detail="没有字段映射规则，请先配置映射")
    dwd_table = first_mapping.target_dwd_table
    cols = _dwd_columns(dwd_table)
    if not cols:
        raise HTTPException(
            status_code=404,
            detail=f"DWD 表 '{dwd_table}' 不存在，请先执行 Raw→DWD",
        )
    return {"columns": cols, "dwd_table": dwd_table, "src_dwd_table": dwd_table}


@router.get("/{ds_id}/raw-columns")
def get_raw_columns(
    ds_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    ds = _get_ds(ds_id, db)
    cols = _raw_columns(ds.target_raw_table)
    if not cols:
        raise HTTPException(
            status_code=404,
            detail=f"Raw 表 '{ds.target_raw_table}' 不存在或无数据列，请先上传或拉取文件",
        )
    return {"columns": cols, "raw_table": ds.target_raw_table}


@router.get("/{ds_id}/mappings", response_model=list[MappingOut])
def list_mappings(
    ds_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    _get_ds(ds_id, db)
    return (
        db.query(FieldMapping)
        .filter(FieldMapping.data_source_id == ds_id)
        .order_by(FieldMapping.sort_order, FieldMapping.id)
        .all()
    )


@router.put("/{ds_id}/mappings", response_model=list[MappingOut])
def bulk_save_mappings(
    ds_id: int,
    body: BulkSaveRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _get_ds(ds_id, db)
    db.query(FieldMapping).filter(FieldMapping.data_source_id == ds_id).delete()
    saved = []
    for i, item in enumerate(body.mappings):
        m = FieldMapping(
            data_source_id=ds_id,
            src_field=item.src_field,
            dst_field=item.dst_field,
            dst_type=item.dst_type,
            default_value=item.default_value or None,
            skip=item.skip,
            sort_order=item.sort_order or i,
            target_dwd_table=body.target_dwd_table,
        )
        db.add(m)
        saved.append(m)
    db.commit()
    for m in saved:
        db.refresh(m)

    # snapshot for config lineage history
    db.add(MappingVersion(
        data_source_id=ds_id,
        snapshot_json={
            "target_dwd_table": body.target_dwd_table,
            "mappings": [
                {
                    "src_field": m.src_field,
                    "dst_field": m.dst_field,
                    "dst_type": m.dst_type.value,
                    "skip": m.skip,
                    "default_value": m.default_value,
                    "sort_order": m.sort_order,
                }
                for m in saved
            ],
        },
        saved_by=current_user.id,
    ))
    db.commit()

    return saved


@router.get("/{ds_id}/mappings/template")
def download_template(
    ds_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    ds = _get_ds(ds_id, db)
    cols = _raw_columns(ds.target_raw_table)
    if not cols:
        raise HTTPException(status_code=404, detail="Raw 表不存在，请先上传数据")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段映射"

    # Header row
    headers = ["src_field", "dst_field", "dst_type", "default_value", "skip"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    # Pre-fill one row per raw column
    for col in cols:
        ws.append([col, col, "string", "", "FALSE"])

    # Column widths
    for i, w in enumerate([20, 20, 12, 15, 8], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Add a note sheet with allowed dst_type values
    ws2 = wb.create_sheet("说明")
    ws2.append(["dst_type 可选值"])
    for t in DST_TYPES:
        ws2.append([t])
    ws2.append([])
    ws2.append(["skip 填 TRUE 表示跳过该字段（不进入 DWD 层）"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mapping_template_{ds.target_raw_table}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/{ds_id}/mappings/import", response_model=list[MappingOut])
async def import_from_excel(
    ds_id: int,
    target_dwd_table: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    _get_ds(ds_id, db)
    suffix = Path(file.filename or "f").suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx/.xls）")

    tmp = tempfile.mktemp(suffix=suffix)
    with open(tmp, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        os.unlink(tmp)

    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    missing = {"src_field", "dst_field", "dst_type"} - set(header)
    if missing:
        raise HTTPException(status_code=400, detail=f"Excel 缺少必要列：{missing}")

    def val(row, name, default=""):
        idx = header.index(name) if name in header else -1
        v = row[idx] if 0 <= idx < len(row) else None
        return str(v).strip() if v is not None else default

    items = []
    for i, row in enumerate(rows[1:], 1):
        src = val(row, "src_field")
        if not src:
            continue
        skip_raw = val(row, "skip", "FALSE").upper()
        items.append(
            MappingItem(
                src_field=src,
                dst_field=val(row, "dst_field") or src,
                dst_type=val(row, "dst_type", "string") or "string",
                default_value=val(row, "default_value") or None,
                skip=skip_raw in ("TRUE", "1", "YES"),
                sort_order=i,
            )
        )

    if not items:
        raise HTTPException(status_code=400, detail="Excel 中没有有效数据行")

    return bulk_save_mappings(
        ds_id,
        BulkSaveRequest(target_dwd_table=target_dwd_table, mappings=items),
        db,
    )
